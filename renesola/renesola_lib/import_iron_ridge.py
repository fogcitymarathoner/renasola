__author__ = 'marc'
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import sys
import json
from datetime import datetime as dt
import redis
from django.conf import settings

from django.shortcuts import render
from django.template import loader, Context
from pricing.crm.models import ProductClassification

from pricing.crm.models import Product
from pricing.crm.models import ProductClassification
from pricing.crm.models import Price as ProductPrice

from renesola_lib.import_iron_ridge_settings import *

import unicodedata

from kombu.log import get_logger
logger = get_logger(__name__)

def build_racking_descriptions_coffee(rows=None):
    if rows == None:
        # called from command
        # pull data from redis cache build during upload
        # satisfies a make target
        r = redis.StrictRedis(host='localhost', port=settings.REDIS_PORT, db=settings.REDIS_DB)
        print r.get('racking_descriptions')
        rows = json.loads(r.get('racking_descriptions'))
    data = {
        'rows': rows,
    }
    context = Context(data)
    t = loader.get_template("pricing/crm/racking_rows.coffee.txt")
    coffee_contents = t.render(context)
    coffee_script_file = os.path.join(settings.BASE_DIR, 'renesola', 'coffee_script', 'discount_calculator',
                                      'locally_built_data', 'racking_descriptions.coffee')

    fo = open(coffee_script_file, 'w')
    fo.write(coffee_contents)
    fo.close()

def import_description_xl(filename):
    """
    test pricing model loading.
    assumes field delimiter = ';'
    text delimiter ='"'

    decodes to "ISO-8859-1" to account for floating 'x' character for dimensions in
    price sheets

    in the future load directly from Excel
    http://stackoverflow.com/questions/9884353/xls-to-csv-convertor


    """
    print 'in import_description_xl'
    r = redis.StrictRedis(host='localhost', port=settings.REDIS_PORT, db=settings.REDIS_DB)
    o = {
        'upload_date': dt.now().strftime("%A, %d. %B %Y %I:%M%p"),
        'message': 'STILL UPLOADING',
    }
    print 'setting racking descriptions in redis'
    print json.dumps(o)
    r.set('rackinglist_meta', json.dumps(o))
    logger.info('opening workbook through %s'%filename)
    wb = load_workbook(filename = filename, use_iterators = True)
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[1])

    i = 0
    count = 0
    rows = []
    for row in ws.iter_rows():

        data_array = []
        for c in row:
            data_array.append(c.value)

        logger.info( data_array)
        if data_array[X10] == None:
            x10 = ''
        else:

            x10 = unicodedata.normalize('NFKD', data_array[X10]).encode('ascii','ignore')
        if data_array[X100] == None:
            x100 = ''
        else:
            x100 = unicodedata.normalize('NFKD', data_array[X100]).encode('ascii','ignore')

        if data_array[X1000] == None:
            x1000 = ''
        else:

            x1000 = unicodedata.normalize('NFKD', data_array[X1000]).encode('ascii','ignore')
        data = {
            'x10':  x10,
            'x100':  x100,
            'x1000':  x1000,
        }
        if count > 0:
            rows.append(data)
        count += 1

    r.set('racking_descriptions', json.dumps(rows))
    build_racking_descriptions_coffee(rows)
    o['message'] = 'wrote racking_descriptions.coffee from import_description_xl'
    r.set('rackinglist_meta', json.dumps(o))
    logger.info('wrote racking_descriptions.coffee from import_description_xl')
    logger.info('Remaking discount_calculator.js')

    o['upload_date'] = dt.now().strftime("%A, %d. %B %Y %I:%M%p")
    o['message'] = 'Remaking static/js/discount_calculator.js'
    r.set('rackinglist_meta', json.dumps(o))
    # pwd = renesola/renesola/kombu_workers
    if os.path.isfile('../apps/pricing/crm/static/js/discount_calculator.js'):
        os.remove('../apps/pricing/crm/static/js/discount_calculator.js')

    if os.path.isfile('../../static/js/discount_calculator.js'):
        os.remove('../../static/js/discount_calculator.js')

    os.system('cd ../../ && make static/js/discount_calculator.js')
    o['upload_date'] = dt.now().strftime("%A, %d. %B %Y %I:%M%p")
    o['message'] = 'FINISHED!!'
    r.set('rackinglist_meta', json.dumps(o))
def handle_uploaded_file(f):
    """
    Write uploaded file down,
    load into database

    :param f:
    :return:
    """
    uploaded_file_target = '/tmp/racking-descriptions.xlsx'
    with open(uploaded_file_target, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
    logger.info('processing racking spreadsheet uploaded to %s'%uploaded_file_target)
    import_description_xl(uploaded_file_target)
