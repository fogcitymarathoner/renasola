__author__ = 'marc'
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import sys
import json
from datetime import datetime as dt
import redis
from django.conf import settings

from django.shortcuts import render
from django.template import loader, Context
from pricing.crm.models import ProductClassification

from pricing.crm.models import Product
from pricing.crm.models import Price as ProductPrice

from renesola_lib.import_xl_settings import *

from renesola_lib.import_cvs import empty_pricelist_tables
import unicodedata
from kombu.log import get_logger

logger = get_logger(__name__)


def build_categories_coffee():
    categories_products = []
    categories = ProductClassification.objects.all()
    for c in categories:
        products = []
        for p in Product.objects.filter(product_classification=c):
            po = {
                'id': p.id,
                'description': p.description
            }
            products.append(po)
        o = {
            'description': c.description,
            'products': products,
        }
        categories_products.append(o)
    data = {
        "category_list": json.dumps(categories_products)
    }

    context = Context(data)
    t = loader.get_template("pricing/crm/categories.coffee.txt")
    coffee_contents = t.render(context)


    coffee_script_file = os.path.join(settings.BASE_DIR, 'renesola', 'coffee_script', 'discount_calculator',
                                      'locally_built_data', 'categories.coffee')

    fo = open(coffee_script_file, 'w')
    fo.write(coffee_contents)
    fo.close()
    msg = 'wrote categories.coffee from import_xl'

    logger.info(msg)

def import_pricelist_xl(filename):
    """
    test pricing model loading.
    assumes field delimiter = ';'
    text delimiter ='"'

    decodes to "ISO-8859-1" to account for floating 'x' character for dimensions in
    price sheets

    in the future load directly from Excel
    http://stackoverflow.com/questions/9884353/xls-to-csv-convertor

    """
    r = redis.StrictRedis(host='localhost', port=settings.REDIS_PORT, db=settings.REDIS_DB)
    # set preliminary status in redis to indicate progress
    #
    # clear pricelist tables
    empty_pricelist_tables()
    wb = load_workbook(filename = filename, use_iterators = True)
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[0])
    o = {
        'upload_date': dt.now().strftime("%A, %d. %B %Y %I:%M%p"),
        'product_count': 0,
        'products': [],
        'message': 'STILL UPLOADING',
        'total_product_count': ws.get_highest_row() - 1
    }
    r.set('pricelist_meta', json.dumps(o))

    logger.info('wrote pricelist_meta')
    i = 0
    product_count = 0
    for row in ws.iter_rows():
        data_array = []
        for c in row:
            data_array.append(c.value)

        if data_array[PNUM] == None:
            data_array[PNUM] = ''
        product_id = unicodedata.normalize('NFKD', data_array[PID]).encode('ascii','ignore')
        product_number_col = data_array[PNUM]
        product_classification_col = data_array[PCLASSIFICATION]
        region_col = data_array[REGION]
        currency_col = data_array[CURRENCY]

        if product_number_col != None:
            product_number = product_number_col.decode("ISO-8859-1")
        else:
            product_number = ''
        if type(currency_col) == 'unicode':
            currency = currency_col.decode("ISO-8859-1")
        else:
            currency = ''

        data = {
            'product_id':  product_id,
            'product_number':  product_number,
            'product_classification':  data_array[PCLASSIFICATION].decode("ISO-8859-1"),
            'region':  data_array[REGION].decode("ISO-8859-1"),
            'currency':  currency,
            'qb1':  data_array[QB1],
            'qb2':  data_array[QB2],
            'qb3':  data_array[QB3],
            'qb4':  data_array[QB4],
            'exw1':  data_array[EXW1],
            'exw2':  data_array[EXW2],
            'exw3':  data_array[EXW3],
            'exw4':  data_array[EXW4],
        }


        logger.info(data)
        if i > 0:
            dj_product_classification, res = ProductClassification.objects.get_or_create(
                description=data['product_classification'])

            dj_product, res = Product.objects.get_or_create(description=data['product_id'],
                                                        book_region=data['region'],
                                                        product_classification=dj_product_classification)

            product_count += 1
            current_product_description = data['product_classification'] + ': ' + data['product_id']
            o['products'].append(current_product_description)
            o['product_count'] = product_count
            current_message = data['product_classification']+': '+data['product_id']
            o['message'] = 'Inserting %s '%current_message
            r.set('pricelist_meta', json.dumps(o))
            if data['qb1'] != None:
                if data['qb1'] > 0:
                    dj_product_price = ProductPrice.objects.get_or_create(description='0',
                                                                      product=dj_product,
                                                                      price=data['exw1'],
                                                                      price_break_quantity=data['qb1'])
            if data['qb2'] != None:
                if data['qb2'] > 0:
                    dj_product_price = ProductPrice.objects.get_or_create(description='1',
                                                                      product=dj_product,
                                                                      price=data['exw2'],
                                                                      price_break_quantity=data['qb2'])
            if data['qb3'] != None:
                if data['qb3'] > 0:
                    dj_product_price = ProductPrice.objects.get_or_create(description='2',
                                                                      product=dj_product,
                                                                      price=data['exw3'],
                                                                      price_break_quantity=data['qb3'])
            if data['qb4'] != None:
                if data['qb4'] > 0:
                    dj_product_price = ProductPrice.objects.get_or_create(description='3',
                                                                      product=dj_product,
                                                                      price=data['exw4'],
                                                                      price_break_quantity=data['qb4'])


        i += 1

    o['message'] = 'Part 2/3: Building Coffeescript'
    r.set('pricelist_meta', json.dumps(o))
    build_categories_coffee()
    o['message'] = 'Part 3/3: Compiling Coffeescript'
    r.set('pricelist_meta', json.dumps(o))

    logger.info('Remaking discount_calculator.js')

    # pwd = renesola/renesola/kombu_workers
    logger.info('removing ../apps/pricing/crm/static/js/discount_calculator.js')
    if os.path.isfile('../apps/pricing/crm/static/js/discount_calculator.js'):
        os.remove('../apps/pricing/crm/static/js/discount_calculator.js')

    logger.info('removing ../../static/js/discount_calculator.js')
    if os.path.isfile('../../static/js/discount_calculator.js'):
        os.remove('../../static/js/discount_calculator.js')

    logger.info('removing ../coffee_script/discount_calculator/discount_calculator.coffee')
    if os.path.isfile('../coffee_script/discount_calculator/discount_calculator.coffee'):
        os.remove('../coffee_script/discount_calculator/discount_calculator.coffee')


    os.system('cd ../../ && make static/js/discount_calculator.js')
    o['upload_date'] = dt.now().strftime("%A, %d. %B %Y %I:%M%p")
    o['product_count'] = product_count
    o['message'] = 'UPLOAD SUCCESSFUL'

    logger.info('UPLOAD SUCCESSFUL')
    r.set('pricelist_meta', json.dumps(o))

def handle_uploaded_file(f):
    """
    Write uploaded file down,
    load into database

    :param f:
    :return:
    """
    uploaded_file_target = '/tmp/pricelist.xlsx'
    with open(uploaded_file_target, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
    import_pricelist_xl(uploaded_file_target)