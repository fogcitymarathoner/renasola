__author__ = 'marc'
from openpyxl import Workbook
from openpyxl import load_workbook
import os

import tenjin
from tenjin.helpers import *
import sys
import json
from datetime import datetime as dt
import redis
from django.conf import settings

from django.shortcuts import render
from django.template import loader, Context


from renesola_lib.import_product_packaging_settings import *

import unicodedata
from kombu.log import get_logger

from renesola.renesola_lib.mongo_models.products_packaging import ProductPackaging
logger = get_logger(__name__)

def build_categories_coffee():
    products = ProductPackaging().all_details_dict_pallet()

    data = {
        'products': products
    }

    ## change to store template cache into memory instead of file system
    tenjin.Engine.cache = tenjin.MemoryCacheStorage()
    engine = tenjin.Engine(cache=tenjin.MemoryCacheStorage(), path=[settings.TENJIN_TEMPLATES_DIR])


    txt = engine.render('js.pyhtml', data)
    local_data_dir = os.path.join(settings.BASE_DIR, 'renesola', 'coffee_script', 'freight_calculator',
                                      'locally_built_data')
    coffee_script_file = os.path.join(local_data_dir, 'categories.coffee')
    if not os.path.exists(local_data_dir):
            os.makedirs(local_data_dir)
    fo = open(coffee_script_file, 'w')
    fo.write(txt)
    fo.close()
    msg = 'wrote categories.coffee from import_xl'

    logger.info(msg)

def import_product_packaging(filename):
    """
    test product packaging loading.
    assumes field delimiter = ';'
    text delimiter ='"'

    decodes to "ISO-8859-1" to account for floating 'x' character for dimensions in
    price sheets

    in the future load directly from Excel
    http://stackoverflow.com/questions/9884353/xls-to-csv-convertor

    """
    # set preliminary status in redis to indicate progress
    #
    # clear pricelist tables

    wb = load_workbook(filename = filename, use_iterators = True)
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[1])
    products = []
    i = 0
    for row in ws.iter_rows():
        if i > 3:
            data_array = []
            for c in row:
                #logger.info(c.value)
                data_array.append(c.value)
            product_category = data_array[PRODUCT_CATEGORY]
            logger.info('product category')
            logger.info(product_category)
            if product_category is not None:
                if product_category != 'Manual Input':
                    unit_spec_length = data_array[UNIT_SPEC_LENGTH]
                    unit_spec_width = data_array[UNIT_SPEC_WIDTH]
                    unit_spec_height = data_array[UNIT_SPEC_HEIGHT]
                    unit_spec_weight = data_array[UNIT_SPEC_WEIGHT]

                    carton_spec_length = data_array[CARTON_SPEC_LENGTH]
                    carton_spec_width = data_array[CARTON_SPEC_WIDTH]
                    carton_spec_height = data_array[CARTON_SPEC_HEIGHT]
                    carton_spec_weight = data_array[CARTON_SPEC_WEIGHT]
                    carton_spec_quantity = data_array[CARTON_SPEC_QUANTITY]

                    pallet_spec_length = data_array[PALLET_SPEC_LENGTH]
                    pallet_spec_width = data_array[PALLET_SPEC_WIDTH]
                    pallet_spec_height = data_array[PALLET_SPEC_HEIGHT]
                    pallet_spec_weight = data_array[PALLET_SPEC_WEIGHT]
                    pallet_spec_quantity = data_array[PALLET_SPEC_QUANTITY]

                    ltl_class = data_array[LTL_CLASS]

                    minimum_packing_unit = data_array[MINIMUM_PACKAGING_UNIT]

                    prod = {
                        'category': product_category,
                        'freight_class': ltl_class,
                        'minimum_packing_unit': minimum_packing_unit,
                        'unit_spec': {
                            'length': unit_spec_length,
                            'width': unit_spec_width,
                            'height': unit_spec_height,
                            'weight': unit_spec_weight
                        },
                        'carton_spec': {
                            'length': carton_spec_length,
                            'width': carton_spec_width,
                            'height': carton_spec_height,
                            'weight': carton_spec_weight,
                            'quantity': carton_spec_quantity
                        },
                        'pallet_spec': {
                            'length': pallet_spec_length,
                            'width': pallet_spec_width,
                            'height': pallet_spec_height,
                            'weight': pallet_spec_weight,
                            'quantity': pallet_spec_quantity
                        }
                    }
                    products.append(prod)
        i += 1
    logger.info('products')
    logger.info(products)

    ProductPackaging().empty_out()

    ProductPackaging().insert(products)